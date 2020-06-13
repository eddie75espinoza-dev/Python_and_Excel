'''
Esta aplicación es una prueba de conexion con un archivo de excel, permite el registro de datos,
tiene un validador de campo en blanco. No se implementó para incorporar varios registros al mismo
archivo excel. Con la librería OPENPYXL se puede hacer la conexion, y trabajar con la mayoría de
las funciones de excel.
'''

from tkinter import *
from openpyxl import *
from tkinter.messagebox import showinfo
from tkinter import filedialog,messagebox

win = Tk()
win.title("Registro") #Nombre de la ventana de registro
img_icon = PhotoImage(file = 'LogoE256.png')
win.call('wm', 'iconphoto', win._w, img_icon)

#Ruta del archivo usado por openpyxl
path = r'.\Python\PyExcel' #Ruta Archivo
dest_filename = r'.\Libro1.xlsx' # Archivo destino

#Del modulo 'openpyxl' activa el libro de excel
wb = Workbook()

'''
Funcion definida para guardar los registros, sin validadores de campo, sin verificar
que existan otros registros en el archivo destino.
'''
def save():
    if len(entry_n.get()) != 0 or len(entry_e.get()) != 0 or len(entry_n.get()) != 0:
        f_name = entry_f.get()
        e_name = entry_e.get()
        num = int(entry_n.get())

        wb = Workbook()
        ws = wb.active      #Activa la hoja

        ws.title = "Equipo" #titulo de hoja
        ws['A1'] = 'Nombre'
        ws['B1'] = 'Equipo'
        ws['C1'] = 'Numero'
        ws['A2'] = f_name
        ws['B2'] = e_name
        ws['C2'] = num

        wb.save(filename = dest_filename)
        showinfo("Guardado", "Registros guardados con exito")
        clear()
    else:
        showinfo("Verificar", "Faltan datos al Registro")
        entry_f.focus()

def clear():
    entry_f.delete(0, END)
    entry_e.delete(0, END)
    entry_n.delete(0, END)
    entry_f.focus()

def open_fold():
    path = filedialog.askdirectory()
    if path == "":
        path = dest_filename

label = Label(win, text="Conexion de Python con Excel", bg= "yellow")
label.grid(row=0, column=0)

f_label = Label(win, text="Nombre: ")
f_label.grid(row=1, column=0, padx=8, pady=8)

e_label = Label(win, text="Equipo: ")
e_label.grid(row=2, column=0, padx=8, pady=8)

n_label = Label(win, text="Número: ")
n_label.grid(row=3, column=0, padx=8, pady=8)

entry_f = Entry(win)
entry_f.grid(row=1, column=1, padx=8, pady=8)
entry_f.focus()

entry_e = Entry(win)
entry_e.grid(row=2, column=1, padx=8, pady=8)

entry_n = Entry(win)
entry_n.grid(row=3, column=1, padx=8, pady=8)

button = Button(win, text="Registrar", command=save)
button.grid(row=5, column=0, columnspan=2, padx=8, pady=8)

button2 = Button(win, text="Limpiar", command=clear)
button2.grid(row=5, column=1, columnspan=2, padx=8, pady=8)

button3 = Button(win,text = "carpeta", command=open_fold)
button3.grid(row=7, column=0, columnspan= 3, padx=8, pady=8)

wb = load_workbook(filename = 'Libro1.xlsx')
sheet_ranges = wb['Equipo']
win.mainloop()

